import json
from openpyxl import load_workbook

path = r"C:\Users\selly\Desktop\app\quest pool.xlsx"
wb = load_workbook(path, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(values_only=True))

# column index -> (category, subcategory key, display label, shape svg name)
COLS = {
    0: ("social",  "game",           "Game",           "game"),
    1: ("social",  "activity",       "Activity",       "activity"),
    2: ("social",  "encounter",      "Encounter",      "encounter"),
    3: ("outside", "discovery",      "Discovery",      "discovery"),
    4: ("outside", "action",         "Action",         "action"),
    5: ("outside", "scavenger_hunt", "Scavenger Hunt", "scavenger_hunt"),
    6: ("indoor",  "tasks",          "Tasks",          "tasks"),
    7: ("indoor",  "diy",            "DIY",            "creative"),
    8: ("indoor",  "selfcare",       "Selfcare",       "selfcare"),
}
CATS = ["outside", "indoor", "social"]  # green, yellow, purple
pool = {c: {"color": c, "subcategories": []} for c in CATS}

for col, (cat, key, label, shape) in COLS.items():
    quests = []
    for r in range(2, len(rows)):
        v = rows[r][col]
        if v is not None and str(v).strip() != "":
            quests.append(str(v).strip())
    pool[cat]["subcategories"].append({"key": key, "label": label, "shape": shape, "quests": quests})

out = "// Auto-generated from 'quest pool.xlsx'. Do not edit by hand.\n"
out += "window.QUEST_POOL = " + json.dumps(pool, ensure_ascii=False, indent=2) + ";\n"
with open(r"C:\Users\selly\Desktop\app\code\js\quests.js", "w", encoding="utf-8") as f:
    f.write(out)
print("wrote quests.js")
for c in CATS:
    print(c, "->", ", ".join(s["key"] + "(" + str(len(s["quests"])) + ")" for s in pool[c]["subcategories"]))
