import sys, json
from openpyxl import load_workbook

path = r"C:\Users\selly\Desktop\app\quest pool.xlsx"
wb = load_workbook(path, data_only=True)
for ws in wb.worksheets:
    print("=== SHEET:", ws.title, "dims:", ws.dimensions, "===")
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        # skip fully empty rows
        if any(c is not None and str(c).strip() != "" for c in row):
            print(i, ['' if c is None else str(c) for c in row])
